#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
产品化服务：上传 -> 后台任务 -> 下载
- POST /api/tasks?type=sales        创建销量处理任务并上传文件（请求体=原始 xlsx 字节）
- POST /api/upload?sid=&slot=        暂存 FBA 差异分析所需的单个文件（销量/运营级别/FBA需求）
- POST /api/tasks?type=fba&sid=      链式处理：销量结果 -> 运营级别结果 -> 差异比对（产出 3 份结果）
- POST /api/tasks?type=eu_check&sid= 多余欧洲区销量核对（差异表 + 销量原始，按 MSKU 累计比对）
- POST /api/tasks?type=alloc_check    上传「分配结果」Excel，按 SKU 分组重算 FBA 每周期分配量并比对实际分配量
- GET  /api/tasks/<id>               查询任务状态/行数/错误/预览/汇总
- GET  /api/tasks/<id>/download?which=sales|ops|diff  下载结果 xlsx
- GET  /                          打开处理页面
启动：python server.py [端口]，默认 8000
"""
import http.server, threading, os, json, uuid, time, sys, re, urllib.parse, zipfile, inspect
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
# 端口：优先读环境变量 $PORT（云平台的标配），其次命令行参数，最后默认 8000
PORT = int(os.environ.get('PORT', sys.argv[1] if len(sys.argv) > 1 else 8000))
DEF = (0.6, 0.3, 0.1)
# 销量处理列名默认（可经前端高级配置覆盖）
DEF_COLS = {'store': '店铺', 'asin': 'ASIN', 'd3': '三天日均销量', 'd7': '七天日均销量', 'd14': '十四天日均销量'}
# 运营级别处理列名默认
DEF_OPS_COLS = {'store': '店铺', 'asin': 'ASIN', 'status': '销售状态', 'dtype': '配送类型',
                'asinKey': '店铺ASIN', 'salesKey': '日均销量加权', 'level': '运营级别'}

# 备货规则 6 窗口列名（标准中文列名；某窗列缺失按 0 处理，权重=0 不影响结果）
WIN_COLS = {3: "三天日均销量", 7: "七天日均销量", 14: "十四天日均销量",
            30: "三十天日均销量", 60: "60天日均销量", 90: "90天日均销量"}


def load_mapping(path):
    """读运营级别映射表 -> {店铺+ASIN(去空格拼接): (运营级别, 补货方式)}。
    支持两种列格式：①「店铺」+「ASIN」两列；②「店铺ASIN」合并一列。
    合并 key 与销量结果 A 列（店铺+ASIN 无分隔符）构造方式一致，故两种格式都能匹配。
    """
    try:
        header, it = _calamine_rows(path)
    except Exception:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(h) if h is not None else '' for h in next(it)]
    if not header:
        raise ValueError("映射表没有任何数据行（只有表头或为空）")

    def find(n):
        for i, h in enumerate(header):
            if h == n:
                return i
        return -1
    i_store = find('店铺'); i_asin = find('ASIN'); i_comb = find('店铺ASIN')
    i_level = find('运营级别'); i_mode = find('补货方式')
    if (i_store < 0 or i_asin < 0) and i_comb < 0:
        raise ValueError("映射表缺少必要列：需有「店铺」+「ASIN」两列，或「店铺ASIN」合并列")
    if i_level < 0:
        raise ValueError("映射表缺少「运营级别」列（当前列：%s）" % "、".join(header))
    m = {}
    for r in it:
        if not r:
            continue
        if i_comb >= 0:
            comb = str(r[i_comb]).strip() if r[i_comb] is not None else ''
        else:
            s = str(r[i_store]).strip() if r[i_store] is not None else ''
            a = str(r[i_asin]).strip() if r[i_asin] is not None else ''
            comb = s + a
        if not comb:
            continue
        level = str(r[i_level]).strip() if r[i_level] is not None else ''
        mode = str(r[i_mode]).strip() if (i_mode >= 0 and r[i_mode] is not None) else ''
        m[comb] = (level, mode)
    return m


def _rule_days(rule):
    """推导一条规则的『天数』（生成需求的天数窗口）。
    显式 days 优先；否则 固定窗=fixed_window，组合窗=权重里最大窗口。"""
    d = rule.get('days')
    if isinstance(d, int) and d > 0:
        return d
    if isinstance(d, str) and d.strip().isdigit():
        return int(d)
    if rule.get('type') == '固定':
        fw = int(rule.get('fixed_window') or 0)
        return fw if fw > 0 else 30
    w = {int(k): float(v) for k, v in (rule.get('weights') or {}).items()}
    return max(w) if w else 30


def _build_rule_lookup(cfg):
    """cfg -> (rules_by_name, lookup_by_op_level)，并校验绑定引用的规则名都存在。
    lookup_by_op_level: 运营级别 -> [规则,...]（一个运营级别可绑多条规则）。
    单字符串绑定会自动规整为 [字符串]。
    """
    rules = {r['name']: r for r in cfg.get('rules', [])}
    if not rules:
        raise ValueError("配置缺少 rules（备货规则）")
    dr = cfg.get('default_rule')
    if dr and dr not in rules:
        raise ValueError("default_rule 引用的规则不存在：%s" % dr)
    lookup = {}
    for op_level, rv in cfg.get('bindings', {}).items():
        names = rv if isinstance(rv, (list, tuple)) else [rv]
        names = [n for n in names if n]
        norm = []
        for n in names:
            if n not in rules:
                raise ValueError("绑定表引用的规则不存在：运营级别「%s」→ 规则「%s」" % (op_level, n))
            norm.append(rules[n])
        if norm:
            lookup[op_level] = norm
    return rules, lookup


TASKS = {}      # id -> dict
UPLOADS = {}    # sid -> {slot: path}
LOCK = threading.Lock()


def fnum(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def col_letter(idx):
    # 0-based -> Excel 列字母（A, B, ... Z, AA, ...）
    s = ''; idx += 1
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _write_xlsx(out_path, sheet_tmp, sheet_name='sheet1'):
    """把流式生成的 sheet XML 打包成最小可用的 xlsx（inline strings，无共享字符串表）。"""
    ct = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
          '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
          '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
          '<Default Extension="xml" ContentType="application/xml"/>'
          '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
          '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
          '</Types>')
    rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '</Relationships>')
    wb = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
          '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
          'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
          '<sheets><sheet name="%s" sheetId="1" r:id="rId1"/></sheets></workbook>' % sheet_name)
    wbres = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
             '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
             '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
             '</Relationships>')
    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', ct)
        z.writestr('_rels/.rels', rels)
        z.writestr('xl/workbook.xml', wb)
        z.writestr('xl/_rels/workbook.xml.rels', wbres)
        # 工作表 XML 占文件体积 99%，用压缩等级 1（最快）：对高度重复的 XML 仍
        # 能压到 ~1.3x，但比默认等级 6 快 4~5 倍。小文件用默认即可。
        z.write(sheet_tmp, 'xl/worksheets/sheet1.xml', compresslevel=1)


def _sheet_tmp():
    return os.path.join(HERE, '_sheet_%d.tmp' % os.getpid())


def _stream_sheet(out_tmp, header, row_iter):
    """把表头 + 逐行数据流式拼成 worksheet XML（inline strings，无共享字符串表）。"""
    from xml.sax.saxutils import escape as _xesc
    def esc(s):
        return _xesc(str(s), {'"': '&quot;'})
    with open(out_tmp, 'w', encoding='utf-8', newline='') as f:
        f.write('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n')
        f.write('<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>')
        def write_row(rownum, cells):
            parts = ['<row r="%d">' % rownum]
            for ci, v in enumerate(cells):
                ref = col_letter(ci) + str(rownum)
                if v is None or v == '':
                    parts.append('<c r="%s"/>' % ref)
                elif isinstance(v, bool):
                    parts.append('<c r="%s" t="inlineStr"><is><t>%s</t></is></c>' % (ref, esc(v)))
                elif isinstance(v, float):
                    parts.append('<c r="%s"><v>%s</v></c>' % (ref, repr(v)))
                elif isinstance(v, int):
                    parts.append('<c r="%s"><v>%d</v></c>' % (ref, v))
                else:
                    parts.append('<c r="%s" t="inlineStr"><is><t xml:space="preserve">%s</t></is></c>' % (ref, esc(v)))
            parts.append('</row>')
            f.write(''.join(parts))
        write_row(1, header)
        rownum = 2
        for row in row_iter:
            write_row(rownum, row)
            rownum += 1
        f.write('</sheetData></worksheet>')


def _calamine_rows(path):
    """优先用 calamine（Rust 解析，惰性迭代，内存恒定）；失败回退 openpyxl read_only。"""
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_path(path)
    sheet = wb.get_sheet_by_index(0)
    it = sheet.iter_rows()                      # 惰性迭代器，首行即表头
    header = [str(h) if h is not None else '' for h in next(it)]
    if not header:
        raise ValueError("文件没有任何数据行（只有表头或为空）")
    return header, it


def process_file(in_path, w3, w7, w14, out_path, cols=None, on_progress=None, cfg=None, mapping=None):
    """销量处理：原始销量 -> 店铺ASIN + 加权日均销量（in_memory 计算，流式写出）。
    - cfg=None 或 cfg['mode']=='legacy' → 沿用旧 w3/w7/w14 单一权重（完全向下兼容）。
    - cfg['mode']=='rules' → 按 (店铺,ASIN) 查 mapping 得运营级别 → bindings 找规则（6窗组合/固定窗）算加权日均。
    """
    C = dict(DEF_COLS)
    if cols:
        for k, v in cols.items():
            if v:
                C[k] = v
    use_rules = bool(cfg and isinstance(cfg, dict) and cfg.get('mode') == 'rules')
    WSUM = w3 + w7 + w14
    if not use_rules and WSUM == 0:
        raise ValueError("三个权重不能同时为 0")

    try:
        header, data_iter = _calamine_rows(in_path)
    except Exception:
        wb_r = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
        ws_r = wb_r.active
        it = ws_r.iter_rows(values_only=True)
        header = [str(h) if h is not None else '' for h in next(it)]
        if not header:
            raise ValueError("文件没有任何数据行（只有表头或为空）")
        data_iter = it

    def find(n):
        for i, h in enumerate(header):
            if h == n:
                return i
        return -1
    i_asin = find(C['asin']); i_store = find(C['store'])
    i_d3 = find(C['d3']); i_d7 = find(C['d7']); i_d14 = find(C['d14'])
    # 6 窗口列（含可选 cols 覆盖 d30/d60/d90）
    def win_idx(win):
        ov = (cols or {}).get('d%d' % win) if cols else None
        name = ov or WIN_COLS[win]
        return find(name)
    i_d30 = win_idx(30); i_d60 = win_idx(60); i_d90 = win_idx(90)
    names = [('店铺', C['store']), ('ASIN', C['asin']),
             ('三天日均销量', C['d3']), ('七天日均销量', C['d7']), ('十四天日均销量', C['d14'])]
    miss = [label + '(' + name + ')' for label, name in names if find(name) < 0]
    if miss:
        raise ValueError("缺少必要列：" + "、".join(miss) + "。当前文件的列名为：" + "、".join(header))

    # ---- 规则模式预编译：{规则名: ('fixed', col_idx) | ('combo', [(w,col_idx)...], wsum)} ----
    if use_rules:
        rules_by_name, lookup = _build_rule_lookup(cfg)
        def plan_for(rule):
            if rule.get('type') == '固定':
                fw = int(rule.get('fixed_window') or 0)
                return ('fixed', win_idx(fw) if fw in WIN_COLS else -1)
            w = {int(k): float(v) for k, v in (rule.get('weights') or {}).items()}
            terms = []
            for win, wt in w.items():
                ci = win_idx(win) if win in WIN_COLS else -1
                if ci >= 0:
                    terms.append((wt, ci))
            wsum = sum(w.values()) or 1.0
            return ('combo', terms, wsum)
        rule_plans = {name: plan_for(r) for name, r in rules_by_name.items()}
        default_rule_name = cfg.get('default_rule')
        default_op_level = cfg.get('default_op_level') or ''

    new_header = list(header)
    new_header.insert(0, '店铺ASIN')
    new_header.insert(i_d14 + 2, '日均销量加权')
    hdr_len = len(new_header)

    def get(row, idx):
        if 0 <= idx < len(row):
            return row[idx]
        return None

    preview_rows = []
    n = 0

    def _compute_rules_m(row, store, asin):
        key = (str(store).strip() + str(asin).strip()) if (store is not None or asin is not None) else ''
        ent = mapping.get(key) if mapping else None
        op_level = ent[0] if ent else default_op_level
        rules_list = lookup.get(op_level)
        if not rules_list:
            rules_list = rules_by_name.get(default_rule_name)
        # 一个运营级别可绑多条规则：Tab① 销量结果每行仍输出单条加权日均，
        # 取首条绑定规则（或默认规则）作为代表值；按天数展开在 process_ops 完成。
        rule = rules_list[0] if rules_list else None
        if not rule:
            return 0.0
        plan = rule_plans[rule['name']]
        if plan[0] == 'fixed':
            ci = plan[1]
            return fnum(get(row, ci)) if ci >= 0 else 0.0
        _, terms, wsum = plan
        val = 0.0
        for wt, ci in terms:
            val += wt * fnum(get(row, ci))
        return val / wsum

    def row_gen():
        nonlocal n
        for row in data_iter:
            store = get(row, i_store); asin = get(row, i_asin)
            a = (str(store) if store is not None else '') + (str(asin) if asin is not None else '')
            if use_rules:
                m = _compute_rules_m(row, store, asin)
            else:
                m = (w3 * fnum(get(row, i_d3)) + w7 * fnum(get(row, i_d7)) + w14 * fnum(get(row, i_d14))) / WSUM
            m = round(m * 10000) / 10000
            nr = list(row); nr.insert(0, a); nr.insert(i_d14 + 2, m)
            while len(nr) < hdr_len:
                nr.append('')
            if len(preview_rows) < 200:
                preview_rows.append(nr)
            yield nr
            n += 1
            if on_progress and n % 5000 == 0:
                on_progress(n)

    tmp = _sheet_tmp()
    _stream_sheet(tmp, new_header, row_gen())
    _write_xlsx(out_path, tmp, 'sheet1')
    try:
        os.remove(tmp)
    except OSError:
        pass
    if on_progress:
        on_progress(n)
    return new_header, n, preview_rows


def process_ops(in_ops, in_sales, out_path, cols=None, on_progress=None, levels=None, cfg=None, mapping=None):
    """运营级别处理：按店铺ASIN 从销量结果取加权日均销量(L)，标记满足条件。
    输出表头=[店铺ASIN]+原表头+[销量,满足条件]（legacy）；
    rules 模式额外展开 [规则,天数,满足条件,备注]，一个运营级别绑多条规则 -> 一店铺ASIN 多行(按天数)。

    参数：
      cfg     : 备货规则配置（mode=='rules' 时生效，按绑定规则展开/算加权日均）
      mapping : 运营级别映射表 {店铺ASIN: (运营级别, 补货方式)}（同 server.load_mapping 产物）
    """
    C = dict(DEF_OPS_COLS)
    if cols:
        for k, v in cols.items():
            if v:
                C[k] = v
    use_rules = bool(cfg and isinstance(cfg, dict) and cfg.get('mode') == 'rules')

    def fidx(hdr, name):
        for i, h in enumerate(hdr):
            if h == name:
                return i
        return -1

    # 1) 读销量结果：legacy 聚合成 {店铺ASIN: 加权日均}；rules 模式保留逐 MSKU 的 6 窗口原始行
    sh, sit = _calamine_rows(in_sales) if _try_calamine(in_sales) else _openpyxl_rows(in_sales)
    iA = fidx(sh, C['asinKey']); iM = fidx(sh, C['salesKey'])
    iStat = fidx(sh, '状态')   # 销量表的 状态 列（正常销售/停止销售）
    if iA < 0 or iM < 0:
        raise ValueError("销量结果缺少必要列：店铺ASIN / 日均销量加权（当前列：%s）" % sh)

    # rules 模式预编译：规则计划(含 days) + 绑定树(op_level -> [规则,...])
    rule_plans = {}; lookup = {}; default_rule_name = ''; default_op_level = ''
    win_col_idx = {}
    if use_rules:
        def win_col_idx_for(win):
            ov = (cols or {}).get('d%d' % win) if cols else None
            name = ov or WIN_COLS[win]
            return fidx(sh, name)
        win_col_idx = {w: win_col_idx_for(w) for w in WIN_COLS}
        rules_by_name, lookup = _build_rule_lookup(cfg)
        default_rule_name = cfg.get('default_rule') or ''
        default_op_level = cfg.get('default_op_level') or ''

        def plan_for(rule):
            if rule.get('type') == '固定':
                fw = int(rule.get('fixed_window') or 0)
                return ('fixed', win_col_idx.get(fw, -1) if fw in WIN_COLS else -1, _rule_days(rule))
            w = {int(k): float(v) for k, v in (rule.get('weights') or {}).items()}
            terms = []
            for win, wt in w.items():
                ci = win_col_idx.get(win, -1) if win in WIN_COLS else -1
                if ci >= 0:
                    terms.append((wt, ci))
            wsum = sum(w.values()) or 1.0
            return ('combo', terms, wsum, _rule_days(rule))
        rule_plans = {name: plan_for(r) for name, r in rules_by_name.items()}

    def rule_m(plan, row):
        if plan[0] == '固定' or (plan[0] == 'fixed'):
            ci = plan[1]
            return fnum(row[ci]) if (0 <= ci < len(row)) else 0.0
        # combo: ('combo', terms, wsum, days) — 用索引取 terms/wsum，兼容有无 days
        terms = plan[1]
        wsum = plan[2]
        val = 0.0
        for wt, ci in terms:
            if 0 <= ci < len(row):
                val += wt * fnum(row[ci])
        return val / wsum

    if use_rules:
        # 按 店铺ASIN 分组保留「正常销售」MSKU 的原始行（后续按各自规则重算加权日均）
        sales_detail = {}
        for r in sit:
            a = r[iA]
            if a is None:
                continue
            a = str(a)
            if iStat >= 0 and r[iStat] is not None and str(r[iStat]).strip() != '正常销售':
                continue
            sales_detail.setdefault(a, []).append(r)
    else:
        sales = {}
        for r in sit:
            a = r[iA]
            if a is None:
                continue
            a = str(a)
            if iStat >= 0 and r[iStat] is not None and str(r[iStat]).strip() != '正常销售':
                continue
            sales[a] = sales.get(a, 0.0) + fnum(r[iM])

    # 2) 读运营级别
    oh, oit = _calamine_rows(in_ops) if _try_calamine(in_ops) else _openpyxl_rows(in_ops)
    i_store = fidx(oh, C['store']); i_asin = fidx(oh, C['asin'])
    i_status = fidx(oh, C['status']); i_dtype = fidx(oh, C['dtype'])
    i_key = fidx(oh, C['asinKey'])
    if i_status < 0 or i_dtype < 0:
        raise ValueError("运营级别缺少必要列：销售状态 / 配送类型（当前列：%s）" % oh)
    i_level = fidx(oh, '运营级别')
    # 运营级别选择：legacy 用 levels 白名单；rules 模式用 binding(lookup 即"被选中"的运营级别)
    allowed = set(lookup.keys()) if use_rules else set()
    if not use_rules and levels:
        items = levels if isinstance(levels, (list, tuple, set)) else [levels]
        for lv in items:
            if not lv:
                continue
            for part in re.split(r'[,\s;；、]+', str(lv).strip()):
                part = part.strip()
                if part:
                    allowed.add(part)
    if (not use_rules) and allowed and i_level < 0:
        raise ValueError("运营级别缺少必要列：运营级别（当前列：%s）" % oh)
    distinct_levels = set()

    # 清洗表头：去掉名为空的列；若原表已有 店铺ASIN/销量 则不重复添加
    clean_idx = [i for i, h in enumerate(oh) if h != '']
    clean_oh = [oh[i] for i in clean_idx]
    i_key2 = fidx(clean_oh, '店铺ASIN')
    iL2 = fidx(clean_oh, '销量')
    prepend_a = (i_key2 < 0)

    if use_rules:
        out_header = list(clean_oh)
        if prepend_a:
            out_header = ['店铺ASIN'] + out_header
        if iL2 < 0:
            out_header = out_header + ['销量', '规则', '天数', '满足条件', '备注']
        else:
            out_header = out_header + ['规则', '天数', '满足条件', '备注']
    else:
        out_header = list(clean_oh)
        if prepend_a:
            out_header = ['店铺ASIN'] + out_header
        if iL2 < 0:
            out_header = out_header + ['销量', '满足条件']
        else:
            out_header = out_header + ['满足条件']
    sales_idx = (iL2 + (1 if prepend_a else 0)) if iL2 >= 0 else None

    n = 0
    meet = 0

    def row_gen():
        nonlocal n, meet
        for r in oit:
            store = r[i_store] if i_store >= 0 else ''
            asin = r[i_asin] if i_asin >= 0 else ''
            a = (str(store) if store is not None else '') + (str(asin) if asin is not None else '')
            if i_key >= 0 and r[i_key] is not None:
                a = str(r[i_key])
            level_val = str(r[i_level]).strip() if (i_level >= 0 and r[i_level] is not None) else ''
            if level_val:
                distinct_levels.add(level_val)

            cr_base = [r[i] for i in clean_idx]
            if prepend_a:
                cr_base = [a] + cr_base

            if not use_rules:
                # ---- legacy：单行 + 白名单 ----
                L = sales.get(a, 0.0)
                level_ok = (not allowed) or (level_val in allowed)
                ok = (str(r[i_status]).strip() == '正常销售') and (str(r[i_dtype]).strip() == 'FBA') and L > 0 and level_ok
                if ok:
                    meet += 1
                if iL2 < 0:
                    cr = cr_base + [round(L * 10000) / 10000]
                else:
                    cr_base[sales_idx] = round(L * 10000) / 10000
                    cr = cr_base
                cr = cr + ['是' if ok else '否']
                yield cr
                n += 1
                if on_progress and n % 5000 == 0:
                    on_progress(n)
                continue

            # ---- rules 模式：按绑定规则逐条展开 ----
            op_level = (mapping.get(a) or (default_op_level, ''))[0] if mapping else default_op_level
            rules_list = lookup.get(op_level)   # list 或 None
            status_ok = (str(r[i_status]).strip() == '正常销售') if i_status >= 0 else True
            dtype_ok = (str(r[i_dtype]).strip() == 'FBA') if i_dtype >= 0 else True
            if not rules_list:
                # 未绑定规则 / 运营级别未选 -> 单行，不满足条件
                if iL2 < 0:
                    cr = cr_base + [0, '', '', '否', '未绑定规则(运营级别未选)']
                else:
                    cr_base[sales_idx] = 0
                    cr = cr_base + ['', '', '否', '未绑定规则(运营级别未选)']
                yield cr
                n += 1
                if on_progress and n % 5000 == 0:
                    on_progress(n)
                continue
            for rule in rules_list:
                plan = rule_plans[rule['name']]
                days = plan[2] if plan[0] == 'fixed' else plan[3]
                M_rule = sum(rule_m(plan, row) for row in sales_detail.get(a, []))
                M_rule = round(M_rule * 10000) / 10000
                m_ok = M_rule > 0
                if status_ok and dtype_ok and m_ok:
                    ok = True
                    note = ''
                else:
                    ok = False
                    if not status_ok:
                        note = '非正常销售'
                    elif not dtype_ok:
                        note = '非FBA'
                    elif not m_ok:
                        note = '该规则加权日均=0'
                    else:
                        note = ''
                if ok:
                    meet += 1
                if iL2 < 0:
                    cr = cr_base + [M_rule, rule['name'], days, '是' if ok else '否', note]
                else:
                    tmp = list(cr_base)
                    tmp[sales_idx] = M_rule
                    cr = tmp + [rule['name'], days, '是' if ok else '否', note]
                yield cr
                n += 1
                if on_progress and n % 5000 == 0:
                    on_progress(n)

    tmp = _sheet_tmp()
    _stream_sheet(tmp, out_header, row_gen())
    _write_xlsx(out_path, tmp, 'sheet1')
    try:
        os.remove(tmp)
    except OSError:
        pass
    if on_progress:
        on_progress(n)
    level_info = {'distinct': sorted(distinct_levels), 'allowed': sorted(allowed)}
    return out_header, n, meet, level_info


def process_diff(in_ops_result, in_fba, out_path, cols=None, on_progress=None, preview_cap=200, cfg=None, mapping=None):
    """差异比对：运营级别结果 vs FBA需求。
    - rules 模式（运营级别结果含『天数』列）且 FBA需求 含『备货周期/天数』列时，
      比对键 = (店铺ASIN, 天数)，逐天比对并标注不满足原因。
    - 否则回退按 店铺ASIN 整体比对（legacy 行为，FBA 按店铺ASIN 聚合）。
    输出列：店铺ASIN/差异状态/运营-店铺/运营-ASIN/运营-销量/运营满足条件/规则/天数/FBA需求-行数/
            FBA需求-需求量合计/FBA需求-日均销量(加权)首行/备注。
    """
    def fidx(hdr, name):
        for i, h in enumerate(hdr):
            if h == name:
                return i
        return -1
    def parse_day(v):
        if v is None:
            return None
        try:
            return int(float(v))
        except (ValueError, TypeError):
            pass
        s = str(v).strip()
        dig = re.sub(r'\D', '', s)
        return int(dig) if dig else None

    # 1) 读运营级别结果
    oh, oit = _calamine_rows(in_ops_result) if _try_calamine(in_ops_result) else _openpyxl_rows(in_ops_result)
    iA = fidx(oh, '店铺ASIN'); iStore = fidx(oh, '店铺'); iAsin = fidx(oh, 'ASIN')
    iL = fidx(oh, '销量'); iOk = fidx(oh, '满足条件')
    iRule = fidx(oh, '规则'); iDayOp = fidx(oh, '天数'); iNote = fidx(oh, '备注')
    if iA < 0:
        raise ValueError("运营级别结果缺少 店铺ASIN 列（当前列：%s）" % oh)
    has_op_day = iDayOp >= 0

    S_op = set()
    op_map = {}
    for r in oit:
        a = str(r[iA]) if r[iA] is not None else ''
        if not a:
            continue
        day = parse_day(r[iDayOp]) if has_op_day else None
        ka = (a, day) if has_op_day else a
        ok = (str(r[iOk]).strip() == '是') if iOk >= 0 else False
        if ok:
            S_op.add(ka)
        if ka not in op_map:
            store = r[iStore] if iStore >= 0 else ''
            asin = r[iAsin] if iAsin >= 0 else ''
            L = fnum(r[iL]) if iL >= 0 else 0.0
            rule = str(r[iRule]) if iRule >= 0 else ''
            note = str(r[iNote]) if iNote >= 0 else ''
            op_map[ka] = (str(store) if store is not None else '', str(asin) if asin is not None else '', L, ok, rule, note)

    # 2) 读 FBA需求
    fh, fit = _calamine_rows(in_fba) if _try_calamine(in_fba) else _openpyxl_rows(in_fba)
    jA = fidx(fh, '店铺ASIN'); jStore = fidx(fh, '店铺'); jAsin = fidx(fh, 'ASIN')
    jD = fidx(fh, '日均销量(加权)'); jM = fidx(fh, '需求量')
    # 探测 FBA 的「天数」列（常见命名）
    day_cands = ['备货周期', '天数', '周期', '补货周期', 'FBA备货周期']
    jDay = -1
    for c in day_cands:
        jDay = fidx(fh, c)
        if jDay >= 0:
            break
    has_fba_day = jDay >= 0
    use_day = has_op_day and has_fba_day   # 两侧都有天数才按天比对
    if jA < 0 and (jStore < 0 or jAsin < 0):
        raise ValueError("FBA需求缺少 店铺ASIN 列，或缺少 店铺+ASIN 列（当前列：%s）" % fh)
    compute_key = (jA < 0)
    fba_map = {}
    for r in fit:
        if compute_key:
            a = (str(r[jStore]) if r[jStore] is not None else '') + (str(r[jAsin]) if r[jAsin] is not None else '')
        else:
            a = str(r[jA]) if r[jA] is not None else ''
        if not a:
            continue
        day = parse_day(r[jDay]) if use_day else None
        ka = (a, day) if use_day else a
        d = fnum(r[jD]) if jD >= 0 else 0.0
        q = fnum(r[jM]) if jM >= 0 else 0.0
        fba_map.setdefault(ka, []).append((q, d))

    S_fba = set(fba_map.keys())
    order = {'缺失(运营满足条件、FBA需求未生成)': 0, '多余(FBA需求有、运营不满足条件)': 1, '已生成': 2}

    def status_of(ka):
        in_op = ka in S_op
        in_fba = ka in S_fba
        if in_op and in_fba:
            return '已生成'
        if in_op:
            return '缺失(运营满足条件、FBA需求未生成)'
        return '多余(FBA需求有、运营不满足条件)'

    rows = []
    preview = []
    for ka in (S_op | S_fba):
        st = status_of(ka)
        oc = op_map.get(ka)
        fc = fba_map.get(ka, [])
        a = ka[0] if isinstance(ka, tuple) else ka
        day = ka[1] if isinstance(ka, tuple) else ''
        store = oc[0] if oc else ''
        asin = oc[1] if oc else ''
        L = oc[2] if oc else ''
        ok = oc[3] if oc else ''
        rule = oc[4] if oc else ''
        op_note = oc[5] if oc else ''
        nc = len(fc)
        qsum = round(sum(x[0] for x in fc), 2)
        dfirst = fc[0][1] if fc else ''   # 匹配天数的 FBA 日均销量（按天比对时即该天；保持页签③兼容）
        if st == '已生成':
            note = op_note
        elif use_day:
            if st.startswith('缺失'):
                note = '运营满足条件但FBA未生成该天数需求'
            else:  # 多余
                note = op_note if op_note else 'FBA有该天数需求、运营未生成(未绑定/未选)'
        else:
            note = ''   # legacy：保持原样（空备注）
        row = [a, st, store, asin, L, ('是' if ok else '否') if oc else '', rule, day, nc, qsum, dfirst, note]
        rows.append(row)
        if len(preview) < preview_cap:
            preview.append(row)
    rows.sort(key=lambda x: (order.get(x[1], 9), x[0]))

    miss = sum(1 for x in rows if x[1].startswith('缺失'))
    extra = sum(1 for x in rows if x[1].startswith('多余'))
    both = sum(1 for x in rows if x[1] == '已生成')

    out_header = ['店铺ASIN', '差异状态', '运营-店铺', '运营-ASIN', '运营-销量', '运营满足条件',
                  '规则', '天数', 'FBA需求-行数', 'FBA需求-需求量合计', 'FBA需求-日均销量(加权)首行', '备注']
    tmp = _sheet_tmp()
    _stream_sheet(tmp, out_header, iter(rows))
    _write_xlsx(out_path, tmp, 'sheet1')
    try:
        os.remove(tmp)
    except OSError:
        pass
    summary = {'op_meet': len(S_op), 'fba_unique': len(S_fba),
               'miss': miss, 'extra': extra, 'both': both, 'total': len(rows),
               'use_day': use_day}
    return out_header, len(rows), summary, preview


# 欧洲区国家/站点代码（用于判定店铺是否欧洲区）
EU_CC = {'BE','DE','FR','IT','ES','NL','SE','PL','UK','IE','AT','PT','LU','DK','FI',
         'GR','CZ','RO','HU','BG','SK','SI','HR','LT','LV','EE','CY','MT'}


def _is_european(store, asin):
    """判断 店铺ASIN 是否欧洲区：店铺名含 EU 标识，或国家码为欧洲站点；店铺为空时从 店铺ASIN 反推。"""
    s = (store or '').strip()
    if s:
        tok = s.split('-')
        if 'EU' in tok:
            return True
        cc = tok[3] if len(tok) >= 4 else tok[-1]
        return cc in EU_CC
    a = asin or ''
    if '-EU-' in a:
        return True
    m = re.search(r'(B0[0-9A-Z]{8})$', a)
    sp = a[:m.start()] if m else a
    toks = sp.split('-')
    if 'EU' in toks:
        return True
    return bool(toks) and toks[-1] in EU_CC


def process_eu_check(diff_path, sales_raw_path, out_path, w3, w7, w14, tol=0.01, on_progress=None):
    """多余欧洲区销量核对：差异表『多余』行 × 销量快照(MSKU 级) 加权日均累计。
    输出每行一个 店铺ASIN + 是否欧洲 + 快照MSKU累计 + FBA需求日均销量(加权)首行 + 差值 + 是否一致。
    返回 (header, n, summary, preview)。
    """
    WSUM = w3 + w7 + w14
    if WSUM == 0:
        raise ValueError("三个权重不能同时为 0")

    # 1) 读销量原始(快照)：按 MSKU 维度，把同一 MSKU 在欧洲各店铺的加权日均相加。
    #    关键：欧洲区同一 (ASIN, MSKU) 会跨多个欧洲店铺各有一行（探查确认 67,282 个欧洲 (ASIN,MSKU) 中 58,012 个跨店铺），
    #    所以累计 key 必须是 MSKU（跨欧洲店铺求和），而不是 店铺ASIN（那只算单店铺、会漏算其他欧洲店的同 MSKU 销量）。
    C = dict(DEF_COLS)
    sh, sit = _calamine_rows(sales_raw_path) if _try_calamine(sales_raw_path) else _openpyxl_rows(sales_raw_path)
    def fidx(hdr, name):
        for i, h in enumerate(hdr):
            if h == name:
                return i
        return -1
    i_store = fidx(sh, C['store']); i_asin = fidx(sh, C['asin']); i_msku = fidx(sh, 'MSKU')
    i_d3 = fidx(sh, C['d3']); i_d7 = fidx(sh, C['d7']); i_d14 = fidx(sh, C['d14'])
    i_stat = fidx(sh, '状态')
    if i_store < 0 or i_asin < 0 or i_msku < 0 or i_d3 < 0 or i_d7 < 0 or i_d14 < 0:
        raise ValueError("销量原始缺少必要列：店铺/ASIN/MSKU/三天日均销量/七天日均销量/十四天日均销量（当前列：%s）" % sh)
    msku_snap = {}   # MSKU -> 该 MSKU 在欧洲各店铺的加权日均累计（仅正常销售）
    msku_of = {}     # 店铺ASIN -> set(MSKU)  （欧洲，用于从差异表反查该店铺ASIN对应的 MSKU 集合）
    for r in sit:
        store = r[i_store]; asin = r[i_asin]; msku = r[i_msku]
        if store is None or asin is None or msku is None:
            continue
        if not _is_european(str(store), str(asin)):
            continue
        a = str(store) + str(asin)
        msku_of.setdefault(a, set()).add(str(msku))
        # 仅计「正常销售」MSKU 的加权日均（无 状态 列时汇总全部，避免漏算）
        if i_stat >= 0 and r[i_stat] is not None and str(r[i_stat]).strip() != '正常销售':
            continue
        m = (w3 * fnum(r[i_d3]) + w7 * fnum(r[i_d7]) + w14 * fnum(r[i_d14])) / WSUM
        msku_snap[str(msku)] = msku_snap.get(str(msku), 0.0) + m

    # 2) 读差异表，筛 多余 行
    dh, dit = _calamine_rows(diff_path) if _try_calamine(diff_path) else _openpyxl_rows(diff_path)
    diA = fidx(dh, '店铺ASIN'); diSt = fidx(dh, '运营-店铺'); diFba = fidx(dh, 'FBA需求-日均销量(加权)首行')
    if diA < 0:
        raise ValueError("差异表缺少 店铺ASIN 列（当前列：%s）" % dh)
    if diFba < 0:
        raise ValueError("差异表缺少 FBA需求-日均销量(加权)首行 列（当前列：%s）" % dh)

    def get(r, i):
        return r[i] if (0 <= i < len(r)) else None

    out_header = ['店铺ASIN', '是否欧洲', '快照MSKU累计销量(加权)', 'FBA需求日均销量(加权)首行',
                  '差值(快照-需求)', '是否一致', '快照MSKU数(欧洲跨店铺)', '备注']
    rows = []
    preview = []
    n_total = n_eu = n_noneu = n_consistent = n_inconsistent = n_missing = 0

    for r in dit:
        st = str(get(r, 1)) if get(r, 1) is not None else ''
        if not st.startswith('多余'):
            continue
        n_total += 1
        a = str(get(r, diA)) if get(r, diA) is not None else ''
        store = str(get(r, diSt)) if get(r, diSt) is not None else ''
        fba_val = fnum(get(r, diFba))
        is_eu = _is_european(store, a)
        if is_eu:
            n_eu += 1
        else:
            n_noneu += 1
        msks = msku_of.get(a)
        if msks is None:
            n_missing += 1
            consistent = False
            diff_v = ''
            sv = None
            note = '快照无此店铺ASIN(欧洲)'
        else:
            sv = sum(msku_snap.get(mk, 0.0) for mk in msks)
            if sv == 0:
                consistent = False
                diff_v = round(sv - fba_val, 4)
                note = '快照仅停止销售MSKU(无正常销售)'
            else:
                diff_v = round(sv - fba_val, 4)
                consistent = abs(diff_v) <= tol
                if consistent:
                    n_consistent += 1
                else:
                    n_inconsistent += 1
                note = ''
        row = [a,
               '是' if is_eu else '否',
               round(sv, 4) if sv is not None else '',
               round(fba_val, 4),
               diff_v,
               '是' if consistent else '否',
               len(msks),
               note]
        rows.append(row)
        if len(preview) < 200:
            preview.append(row)

    # 排序：不一致(含缺失)在前、一致在后；同组欧洲在前
    def rank(x):
        consistent = (x[5] == '是')
        eu = (x[1] == '是')
        return (0 if not consistent else 1, 0 if eu else 1, x[0])
    rows.sort(key=rank)

    tmp = _sheet_tmp()
    _stream_sheet(tmp, out_header, iter(rows))
    _write_xlsx(out_path, tmp, 'sheet1')
    try:
        os.remove(tmp)
    except OSError:
        pass
    if on_progress:
        on_progress(n_total)
    summary = {'extra_total': n_total, 'eu_total': n_eu, 'non_eu_total': n_noneu,
               'consistent': n_consistent, 'inconsistent': n_inconsistent, 'snap_missing': n_missing}
    return out_header, n_total, summary, preview


def process_eu_mismatch_detail(diff_path, sales_raw_path, out_path, w3, w7, w14, tol=0.01, on_progress=None):
    """导出页签③中『不一致/缺失』的欧洲区多余行，并展开每个 店铺ASIN 对应的逐 MSKU 明细（含各欧洲店铺的加权日均）。
    输出双 sheet xlsx：sheet1=不一致汇总, sheet2=MSKU明细。返回 (header_summary, n_mismatch, summary)。
    """
    import openpyxl
    WSUM = w3 + w7 + w14
    if WSUM == 0:
        raise ValueError("三个权重不能同时为 0")
    C = dict(DEF_COLS)
    sh, sit = _calamine_rows(sales_raw_path) if _try_calamine(sales_raw_path) else _openpyxl_rows(sales_raw_path)
    def fidx(hdr, name):
        for i, h in enumerate(hdr):
            if h == name:
                return i
        return -1
    i_store = fidx(sh, C['store']); i_asin = fidx(sh, C['asin']); i_msku = fidx(sh, 'MSKU')
    i_d3 = fidx(sh, C['d3']); i_d7 = fidx(sh, C['d7']); i_d14 = fidx(sh, C['d14'])
    i_stat = fidx(sh, '状态')
    if i_store < 0 or i_asin < 0 or i_msku < 0 or i_d3 < 0 or i_d7 < 0 or i_d14 < 0:
        raise ValueError("销量原始缺少必要列：店铺/ASIN/MSKU/三天日均销量/七天日均销量/十四天日均销量（当前列：%s）" % sh)
    msku_snap = {}     # MSKU -> 该 MSKU 在欧洲各店铺的加权日均累计（仅正常销售）
    msku_detail = {}   # 店铺ASIN -> list[(MSKU, 店铺, 状态, d3, d7, d14, m)]  （欧洲，所有状态都保留便于核对）
    for r in sit:
        store = r[i_store]; asin = r[i_asin]; msku = r[i_msku]
        if store is None or asin is None or msku is None:
            continue
        if not _is_european(str(store), str(asin)):
            continue
        a = str(store) + str(asin)
        msku_s = str(msku)
        d3 = fnum(r[i_d3]); d7 = fnum(r[i_d7]); d14 = fnum(r[i_d14])
        m = (w3 * d3 + w7 * d7 + w14 * d14) / WSUM
        stat = str(r[i_stat]).strip() if (i_stat >= 0 and r[i_stat] is not None) else ''
        msku_detail.setdefault(a, []).append((msku_s, str(store), stat, d3, d7, d14, m))
        if stat != '正常销售':
            continue
        msku_snap[msku_s] = msku_snap.get(msku_s, 0.0) + m

    dh, dit = _calamine_rows(diff_path) if _try_calamine(diff_path) else _openpyxl_rows(diff_path)
    diA = fidx(dh, '店铺ASIN'); diSt = fidx(dh, '运营-店铺'); diFba = fidx(dh, 'FBA需求-日均销量(加权)首行')
    if diA < 0:
        raise ValueError("差异表缺少 店铺ASIN 列（当前列：%s）" % dh)
    if diFba < 0:
        raise ValueError("差异表缺少 FBA需求-日均销量(加权)首行 列（当前列：%s）" % dh)
    def get(r, i):
        return r[i] if (0 <= i < len(r)) else None

    hdr_sum = ['店铺ASIN', '是否欧洲', '快照MSKU累计销量(加权)', 'FBA需求日均销量(加权)首行', '差值(快照-需求)', '备注']
    hdr_det = ['店铺ASIN', 'MSKU', '店铺', '状态', '三天日均销量', '七天日均销量', '十四天日均销量', '单MSKU加权日均']
    summary_rows = []
    n_total = n_mismatch = 0
    mismatch_asins = []
    for r in dit:
        st = str(get(r, 1)) if get(r, 1) is not None else ''
        if not st.startswith('多余'):
            continue
        n_total += 1
        a = str(get(r, diA)) if get(r, diA) is not None else ''
        store = str(get(r, diSt)) if get(r, diSt) is not None else ''
        fba_val = fnum(get(r, diFba))
        is_eu = _is_european(store, a)
        msks = msku_detail.get(a)
        if msks is None:
            sv = None; consistent = False; diff_v = ''; note = '快照无此店铺ASIN(欧洲)'
        else:
            sv = sum(msku_snap.get(mk, 0.0) for mk, _s, _st, _d3, _d7, _d14, _m in msks)
            if sv == 0:
                consistent = False; diff_v = round(sv - fba_val, 4); note = '快照仅停止销售MSKU(无正常销售)'
            else:
                diff_v = round(sv - fba_val, 4); consistent = abs(diff_v) <= tol; note = ''
        if not consistent:
            n_mismatch += 1
            mismatch_asins.append(a)
            summary_rows.append([a, '是' if is_eu else '否',
                                  round(sv, 4) if sv is not None else '', round(fba_val, 4), diff_v, note])
        if on_progress:
            on_progress(n_total)
    mismatch_asins = sorted(set(mismatch_asins))
    det_rows = []
    for a in mismatch_asins:
        for (mk, st, stat, d3, d7, d14, m) in msku_detail.get(a, []):
            det_rows.append([a, mk, st, stat, round(d3, 4), round(d7, 4), round(d14, 4), round(m, 4)])

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = '不一致汇总'
    ws1.append(hdr_sum)
    for row in summary_rows:
        ws1.append(row)
    ws2 = wb.create_sheet('MSKU明细')
    ws2.append(hdr_det)
    for row in det_rows:
        ws2.append(row)
    wb.save(out_path)
    summary = {'total_extra': n_total, 'mismatch': n_mismatch, 'detail_rows': len(det_rows)}
    return hdr_sum, n_mismatch, summary


def process_alloc_check(in_path, out_path, on_progress=None, append_types=None):
    """FBA 分配核对：上传「分配结果」Excel，按 SKU 分组、按备货周期从小到大重算每行的应分配量，
    再逐行与文件中的「实际分配量」比对，标记 一致/不一致/缺实际分配量/无法核对(缺SKU)。
    返回 (header, n, summary, preview)。

    分配口径（已用示例 110 行 / 10 SKU 复现 10/10；2026-08-24 加"追加"识别）：
      1) 同 SKU 的库存池 = 本地SKU总可用库存（取整，同 SKU 各地值一致，取最大更稳）。
      2) 周期 0 且 计算方式='需求'（默认按 需求类型 != append_types 识别）：先全分满 = 预计，扣库存池。
      3) 周期 > 0 按升序：R = 剩余库存；有预计分配量的店铺
         raw_i = 去尾取整(R * 销量_i / 总销量)，封顶 ≤ 预计分配量；
         剩余 = R - Σ(去尾封顶值)，按「预计分配量最大」（并列按 剩余容量=预计−已分 最大）
         依次补 min(剩余, 预计−已分)，直到分完。某周期库存能满足该周期全部预期则直接分满。
      4) 周期 0 且 计算方式='追加'：所有其他周期都分完后，按剩余库存、按销量占比同样分配（去尾+封顶+剩余按预计最大）。
         追加行识别：优先读「计算方式」列（值命中 append_types），缺该列时回退读「需求类型」列（方案①）；
         两者都缺则无追加行。列名/白名单均可在前端「追加特征」框调整。
    """
    C = {
        'sku': 'SKU编码', 'inv': '本地SKU总可用库存', 'cyc': 'FBA备货周期',
        'exp': '预计分配量', 'sales': '日均销量(加权)', 'act': '实际分配量',
        'demandType': '需求类型',   # 识别 追加 行的回退列（方案①）
        'calcMode': '计算方式',     # 识别 追加 行的优先列（方案②：系统导出模板加此列后无需改代码）
    }
    # 追加 特征：默认 需求类型='追加'，可由调用方覆盖
    if append_types is None:
        append_types = ['追加']
    append_set = set(s.strip() for s in append_types if s and str(s).strip())

    header, data_iter = (_calamine_rows(in_path) if _try_calamine(in_path) else _openpyxl_rows(in_path))
    def fidx(hdr, name):
        for i, h in enumerate(hdr):
            if h == name:
                return i
        return -1
    iSku = fidx(header, C['sku']); iInv = fidx(header, C['inv'])
    iCyc = fidx(header, C['cyc']); iExp = fidx(header, C['exp'])
    iSales = fidx(header, C['sales']); iAct = fidx(header, C['act'])
    iDemandType = fidx(header, C['demandType'])  # 可选（回退）
    iCalcMode = fidx(header, C['calcMode'])      # 可选（优先）；缺则回退 需求类型
    missing = [C[k] for k, i in (('sku', iSku), ('inv', iInv), ('cyc', iCyc),
                                 ('exp', iExp), ('sales', iSales), ('act', iAct)) if i < 0]
    if missing:
        raise ValueError("分配结果缺少必要列：" + "、".join(missing) + "（当前列：" + "、".join(header) + "）")

    def cap_of(r):
        return int(round(fnum(r[iExp])))

    def is_append_row(r):
        """周期 0 行 是否按 追加 处理（最后才分）。
        优先读『计算方式』列（方案②），缺该列时回退读『需求类型』列（方案①）；两者都缺则无追加行。"""
        if not append_set:
            return False
        # 优先：计算方式 列（系统导出模板加了该列）
        if iCalcMode >= 0 and r[iCalcMode] is not None:
            if str(r[iCalcMode]).strip() in append_set:
                return True
        # 回退：需求类型 列（方案①）
        if iDemandType >= 0 and r[iDemandType] is not None:
            return str(r[iDemandType]).strip() in append_set
        return False

    def allocate_group(rows, inv, result):
        """对一组行（同一周期或追加组）执行分配：全满足→按比例去尾封顶→剩余按预计最大。
        返回更新后的 inv。"""
        crows = [r for r in rows if fnum(r[iExp]) > 0]
        if not crows:
            for r in rows:
                result[id(r)] = 0
            return inv
        sumexp = int(round(sum(fnum(r[iExp]) for r in crows)))
        totsales = sum(fnum(r[iSales]) for r in crows)
        if inv >= sumexp:
            for r in crows:
                result[id(r)] = cap_of(r)
            inv -= sumexp
        else:
            R = int(round(inv))
            alloc = {}
            for r in crows:
                raw = (R * fnum(r[iSales]) / totsales) if totsales > 0 else 0
                alloc[id(r)] = min(int(raw), cap_of(r))     # 去尾取整，封顶 ≤ 预计
            leftover = R - sum(alloc.values())
            order = sorted(crows, key=lambda r: (cap_of(r), cap_of(r) - alloc[id(r)]), reverse=True)
            for r in order:
                if leftover <= 0:
                    break
                give = min(leftover, cap_of(r) - alloc[id(r)])
                if give > 0:
                    alloc[id(r)] += give
                    leftover -= give
            inv = 0
            for r in crows:
                result[id(r)] = alloc[id(r)]
        return inv

    def compute_group(g):
        """g: 同 SKU 的行对象列表。返回 id(row) -> 重算分配量(int)。"""
        result = {}
        if not g:
            return result
        avail = int(round(fnum(g[0][iInv])))
        for r in g:                                  # 同 SKU 各行库存一致，取最大更稳
            avail = max(avail, int(round(fnum(r[iInv]))))
        inv = avail
        # 步骤 1：周期 0 且 非追加（=需求），先全分满 = 预计，扣池
        cyc0_demand = [r for r in g if fnum(r[iCyc]) == 0 and not is_append_row(r)]
        inv -= int(round(sum(fnum(r[iExp]) for r in cyc0_demand)))
        if inv < 0:
            inv = 0
        for r in cyc0_demand:
            result[id(r)] = cap_of(r)
        # 步骤 2：周期 > 0 升序
        cyc_pos = sorted({fnum(r[iCyc]) for r in g if fnum(r[iCyc]) > 0})
        for c in cyc_pos:
            rows = [r for r in g if fnum(r[iCyc]) == c]
            inv = allocate_group(rows, inv, result)
        # 步骤 3：周期 0 且 追加（=追加），最后才分（按销量占比）
        cyc0_append = [r for r in g if fnum(r[iCyc]) == 0 and is_append_row(r)]
        if cyc0_append:
            inv = allocate_group(cyc0_append, inv, result)
        for r in g:
            result.setdefault(id(r), 0)
        return result

    # 读取并分组（缓冲全部行；分配结果文件通常远小于原始销量表，规模可接受）
    groups = {}
    for r in data_iter:
        rr = list(r)
        sku = str(rr[iSku]) if rr[iSku] is not None else ''
        groups.setdefault(sku, []).append(rr)

    out_header = list(header) + ['重算分配量', '实际分配量(原)', '是否一致', '差异']
    preview_rows = []
    n = 0
    consistent = 0
    inconsistent = 0
    nosku = 0
    append_rows_count = 0   # 被识别为 追加 的行数（用于汇总展示）

    def row_gen():
        nonlocal n, consistent, inconsistent, nosku, append_rows_count
        for sku, g in groups.items():
            if not sku:
                # 无 SKU 的行无法分组核对，原样保留并标记
                for r in g:
                    actual_raw = r[iAct]
                    cr_actual = (int(round(fnum(actual_raw))) if (actual_raw is not None and str(actual_raw).strip() != '') else '')
                    cr = list(r) + ['', cr_actual, '无法核对(缺SKU)', '']
                    yield cr
                    n += 1; nosku += 1
                    if len(preview_rows) < 200:
                        preview_rows.append(cr)
                continue
            res = compute_group(g)
            for r in g:
                recomputed = res[id(r)]
                actual_raw = r[iAct]
                if actual_raw is None or str(actual_raw).strip() == '':
                    cr_actual = ''; ok = '缺实际分配量'; cr_diff = ''
                else:
                    cr_actual = int(round(fnum(actual_raw)))
                    cr_diff = recomputed - cr_actual
                    if cr_diff == 0:
                        ok = '一致'; consistent += 1
                    else:
                        ok = '不一致'; inconsistent += 1
                if is_append_row(r):
                    append_rows_count += 1
                cr = list(r) + [recomputed, cr_actual, ok, cr_diff]
                yield cr
                n += 1
                if len(preview_rows) < 200:
                    preview_rows.append(cr)
            if on_progress and n % 5000 == 0:
                on_progress(n)

    tmp = _sheet_tmp()
    _stream_sheet(tmp, out_header, row_gen())
    _write_xlsx(out_path, tmp, 'sheet1')
    try:
        os.remove(tmp)
    except OSError:
        pass
    if on_progress:
        on_progress(n)
    summary = {
        'total': n, 'consistent': consistent, 'inconsistent': inconsistent, 'nosku': nosku,
        'append_rows': append_rows_count,
        'append_types': sorted(append_set) if append_set else [],
    }
    return out_header, n, summary, preview_rows


def _try_calamine(path):
    try:
        from python_calamine import CalamineWorkbook
        CalamineWorkbook.from_path(path)
        return True
    except Exception:
        return False


def _openpyxl_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else '' for h in next(it)]
    if not header:
        raise ValueError("文件没有任何数据行（只有表头或为空）")
    return header, it


class H(http.server.BaseHTTPRequestHandler):
    def log_message(self, *a):
        pass  # 静默

    def _send(self, code, body, ctype, extra=None):
        self.send_response(code)
        self.send_header('Content-Type', ctype)
        self.send_header('Access-Control-Allow-Origin', '*')
        if extra:
            for k, v in extra.items():
                self.send_header(k, v)
        self.end_headers()
        if isinstance(body, (bytes, bytearray)):
            self.wfile.write(body)
        else:
            self.wfile.write(body.encode('utf-8'))

    def _json(self, code, d):
        self._send(code, json.dumps(d, ensure_ascii=False), 'application/json')

    def do_OPTIONS(self):
        # CORS 预检：处理跨源 POST（file:// 或非同源页面访问时浏览器会先发 OPTIONS）
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', '*')
        self.send_header('Access-Control-Max-Age', '86400')
        self.end_headers()

    def do_GET(self):
        p = urllib.parse.urlparse(self.path)
        if p.path == '/api/ping':
            self._json(200, {'ok': True, 'name': 'store-sales-processor'})
            return
        if p.path in ('/', '/index.html'):
            try:
                data = open(os.path.join(HERE, 'store-sales-processor.html'), 'rb').read()
                self._send(200, data, 'text/html; charset=utf-8')
            except FileNotFoundError:
                self._send(404, '页面文件 store-sales-processor.html 未找到', 'text/plain; charset=utf-8')
            return
        m = re.match(r'^/api/tasks/([^/]+)/download$', p.path)
        if m:
            tid = m.group(1)
            which = qs_get(p, 'which') or ''
            with LOCK:
                t = TASKS.get(tid)
            if not t:
                self._json(404, {'error': 'task not found'})
                return
            if which and t.get('outputs'):
                path = t['outputs'].get(which)
            else:
                path = t.get('output_path')
            if not path or not os.path.exists(path):
                self._send(404, 'result not ready', 'text/plain')
                return
            data = open(path, 'rb').read()
            fname = (t.get('download_name')
                     or {'sales': 'sales_result.xlsx', 'ops': 'ops_result.xlsx', 'diff': 'diff_result.xlsx'}.get(which, 'store_sales_result.xlsx'))
            disp = "attachment; filename=\"%s\"; filename*=UTF-8''%s" % (fname, fname)
            self._send(200, data,
                       'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       {'Content-Disposition': disp})
            return
        m = re.match(r'^/api/tasks/([^/]+)$', p.path)
        if m:
            tid = m.group(1)
            with LOCK:
                t = TASKS.get(tid)
            if not t:
                self._json(404, {'error': 'task not found'})
                return
            self._json(200, {
                'id': tid,
                'status': t['status'],
                'filename': t.get('filename'),
                'n_rows': t.get('n_rows'),
                'header': t.get('header'),
                'preview_rows': t.get('preview_rows'),
                'error_msg': t.get('error_msg'),
                'type': t.get('type'),
                'summary': t.get('summary'),
                'has_output': bool(t.get('output_path') and os.path.exists(t['output_path'])),
                'has_outputs': bool(t.get('outputs')),
                'diff_header': t.get('diff_header'),
                'diff_preview': t.get('diff_preview'),
                'level_info': t.get('level_info'),
                'append_types': t.get('append_types'),
                'created': t.get('created'),
                'finished': t.get('finished'),
            })
            return
        if p.path == '/api/tasks':
            with LOCK:
                lst = [{'id': tid, 'status': t['status'], 'filename': t.get('filename'),
                        'n_rows': t.get('n_rows'), 'type': t.get('type')} for tid, t in TASKS.items()]
            self._json(200, {'tasks': lst})
            return
        self._send(404, 'not found', 'text/plain')

    def do_POST(self):
        p = urllib.parse.urlparse(self.path)
        qs = urllib.parse.parse_qs(p.query)

        # ---- 文件暂存（FBA 差异分析需要多文件） ----
        if p.path == '/api/upload':
            sid = qs_get(p, 'sid') or ''
            slot = qs_get(p, 'slot') or ''
            try:
                length = int(self.headers.get('Content-Length', 0))
            except ValueError:
                length = 0
            data = self.rfile.read(length) if length else b''
            if not sid or not slot:
                self._json(400, {'error': 'need sid and slot'})
                return
            d = UPLOADS.setdefault(sid, {})
            path = os.path.join(HERE, '_up_%s_%s.xlsx' % (sid, slot))
            open(path, 'wb').write(data)
            d[slot] = path
            self._json(200, {'ok': True, 'slot': slot, 'size': len(data)})
            return

        if p.path != '/api/tasks':
            self._send(404, 'not found', 'text/plain')
            return

        # ---- FBA 差异分析（链式：销量->运营级别->差异） ----
        typ = qs_get(p, 'type') or 'sales'
        if typ == 'fba':
            sid = qs_get(p, 'sid') or ''
            try:
                w3 = float(qs_get(p, 'w3') or DEF[0])
                w7 = float(qs_get(p, 'w7') or DEF[1])
                w14 = float(qs_get(p, 'w14') or DEF[2])
            except ValueError:
                self._json(400, {'error': 'invalid weights'})
                return
            cfg = _parse_cfg_param(p)
            if cfg is False:
                self._json(400, {'error': 'cfg 不是合法 JSON'})
                return
            up = UPLOADS.get(sid, {})
            if not (up.get('sales') and up.get('ops') and up.get('fba')):
                self._json(400, {'error': '请先上传 销量 / 运营级别 / FBA需求 三个文件'})
                return
            cols = {
                'store': (qs_get(p, 'colStore') or '').strip(),
                'asin': (qs_get(p, 'colAsin') or '').strip(),
                'd3': (qs_get(p, 'colD3') or '').strip(),
                'd7': (qs_get(p, 'colD7') or '').strip(),
                'd14': (qs_get(p, 'colD14') or '').strip(),
            }
            if not any(cols.values()):
                cols = None
            # 符合补货条件的运营级别（白名单），可重复传多个 levels= 参数
            levels = []
            for raw in qs_get_list(p, 'levels'):
                for part in re.split(r'[,\s;；、]+', raw):
                    part = part.strip()
                    if part:
                        levels.append(part)
            tid = uuid.uuid4().hex[:12]
            task = {
                'id': tid, 'status': 'running', 'type': 'fba', 'filename': 'fba',
                'n_rows': 0, 'header': None, 'preview_rows': None,
                'outputs': None, 'summary': None, 'diff_header': None, 'diff_preview': None,
                'error_msg': None, 'w': (w3, w7, w14), 'cols': cols, 'levels': levels, 'sid': sid,
                'cfg': cfg,
                'created': time.time(), 'finished': None,
            }
            with LOCK:
                TASKS[tid] = task
            threading.Thread(target=run_fba, args=(sid, task), daemon=True).start()
            self._json(202, {'id': tid, 'status': 'running', 'n_rows': 0})
            return

        # ---- 多余欧洲区销量核对（差异表 + 销量原始，MSKU 累计比对） ----
        if typ == 'eu_check':
            sid = qs_get(p, 'sid') or ''
            try:
                w3 = float(qs_get(p, 'w3') or DEF[0])
                w7 = float(qs_get(p, 'w7') or DEF[1])
                w14 = float(qs_get(p, 'w14') or DEF[2])
            except ValueError:
                self._json(400, {'error': 'invalid weights'})
                return
            up = UPLOADS.get(sid, {})
            if not (up.get('diff') and up.get('sales')):
                self._json(400, {'error': '请先上传 差异表 与 销量原始文件'})
                return
            tid = uuid.uuid4().hex[:12]
            task = {
                'id': tid, 'status': 'running', 'type': 'eu_check', 'filename': 'eu_check',
                'n_rows': 0, 'header': None, 'preview_rows': None, 'summary': None,
                'error_msg': None, 'w': (w3, w7, w14), 'sid': sid,
                'download_name': 'eu_check_result.xlsx',
                'created': time.time(), 'finished': None,
            }
            with LOCK:
                TASKS[tid] = task
            threading.Thread(target=run_eu_check, args=(sid, task), daemon=True).start()
            self._json(202, {'id': tid, 'status': 'running', 'n_rows': 0})
            return

        # ---- 多余欧洲区销量核对：导出不一致明细(含逐MSKU) ----
        if typ == 'eu_detail':
            sid = qs_get(p, 'sid') or ''
            try:
                w3 = float(qs_get(p, 'w3') or DEF[0])
                w7 = float(qs_get(p, 'w7') or DEF[1])
                w14 = float(qs_get(p, 'w14') or DEF[2])
            except ValueError:
                self._json(400, {'error': 'invalid weights'})
                return
            up = UPLOADS.get(sid, {})
            if not (up.get('diff') and up.get('sales')):
                self._json(400, {'error': '请先上传 差异表 与 销量原始文件'})
                return
            tid = uuid.uuid4().hex[:12]
            task = {
                'id': tid, 'status': 'running', 'type': 'eu_detail', 'filename': 'eu_detail',
                'n_rows': 0, 'header': None, 'preview_rows': None, 'summary': None,
                'error_msg': None, 'w': (w3, w7, w14), 'sid': sid,
                'download_name': 'eu_mismatch_detail.xlsx',
                'created': time.time(), 'finished': None,
            }
            with LOCK:
                TASKS[tid] = task
            threading.Thread(target=run_eu_detail, args=(sid, task), daemon=True).start()
            self._json(202, {'id': tid, 'status': 'running', 'n_rows': 0})
            return

        # ---- FBA 分配核对（单文件：分配结果） ----
        if typ == 'alloc_check':
            try:
                length = int(self.headers.get('Content-Length', 0))
            except ValueError:
                length = 0
            body = self.rfile.read(length) if length else b''
            fname = (qs_get(p, 'filename') or self.headers.get('X-Filename') or 'alloc.xlsx')
            # 追加特征：可重复传 appendTypes=追加&appendTypes=...，内部再按 [,\s;；、]+ 拆分
            append_types_raw = []
            for raw in qs_get_list(p, 'appendTypes'):
                for part in re.split(r'[,\s;；、]+', raw):
                    part = part.strip()
                    if part:
                        append_types_raw.append(part)
            tid = uuid.uuid4().hex[:12]
            in_path = os.path.join(HERE, '_allocin_' + tid + '.xlsx')
            out_path = os.path.join(HERE, '_allocout_' + tid + '.xlsx')
            open(in_path, 'wb').write(body)
            task = {
                'id': tid, 'status': 'running', 'type': 'alloc_check', 'filename': fname,
                'n_rows': 0, 'header': None, 'preview_rows': None, 'summary': None,
                'error_msg': None, 'input_path': in_path, 'output_path': out_path,
                'download_name': 'alloc_check_result.xlsx',
                'append_types': append_types_raw,   # 透传，便于前端回显
                'created': time.time(), 'finished': None,
            }
            with LOCK:
                TASKS[tid] = task

            def run():
                try:
                    def prog(n):
                        with LOCK:
                            task['n_rows'] = n
                    h, n, summary, prev = process_alloc_check(
                        in_path, out_path, on_progress=prog,
                        append_types=(append_types_raw or None))  # 空列表=用默认 ['追加']
                    with LOCK:
                        task['status'] = 'done'
                        task['n_rows'] = n
                        task['header'] = h
                        task['preview_rows'] = prev
                        task['summary'] = summary
                        task['finished'] = time.time()
                    try:
                        os.unlink(in_path)
                    except OSError:
                        pass
                except Exception as e:
                    import traceback
                    with LOCK:
                        task['status'] = 'error'
                        task['error_msg'] = str(e)
                        task['traceback'] = traceback.format_exc()
                        task['finished'] = time.time()
            threading.Thread(target=run, daemon=True).start()
            self._json(202, {'id': tid, 'status': 'running', 'n_rows': 0})
            return

        # ---- 销量处理（单文件） ----
        try:
            w3 = float(qs_get(p, 'w3') or DEF[0])
            w7 = float(qs_get(p, 'w7') or DEF[1])
            w14 = float(qs_get(p, 'w14') or DEF[2])
        except ValueError:
            self._json(400, {'error': 'invalid weights'})
            return
        # 可选：运营级别映射表（slot=mapping，同 sid）+ 备货规则配置 cfg(JSON)
        sid = qs_get(p, 'sid') or ''
        cfg = _parse_cfg_param(p)
        if cfg is False:
            self._json(400, {'error': 'cfg 不是合法 JSON'})
            return
        try:
            length = int(self.headers.get('Content-Length', 0))
        except ValueError:
            length = 0
        body = self.rfile.read(length) if length else b''
        fname = (qs_get(p, 'filename') or self.headers.get('X-Filename') or 'upload.xlsx')
        cols = {
            'store': (qs_get(p, 'colStore') or '').strip(),
            'asin': (qs_get(p, 'colAsin') or '').strip(),
            'd3': (qs_get(p, 'colD3') or '').strip(),
            'd7': (qs_get(p, 'colD7') or '').strip(),
            'd14': (qs_get(p, 'colD14') or '').strip(),
        }
        if not any(cols.values()):
            cols = None
        tid = uuid.uuid4().hex[:12]
        in_path = os.path.join(HERE, '_in_' + tid + '.xlsx')
        out_path = os.path.join(HERE, '_out_' + tid + '.xlsx')
        open(in_path, 'wb').write(body)
        task = {
            'id': tid, 'status': 'running', 'type': 'sales', 'filename': fname, 'n_rows': 0,
            'header': None, 'preview_rows': None, 'error_msg': None,
            'input_path': in_path, 'output_path': out_path,
            'w': (w3, w7, w14), 'cols': cols, 'sid': sid, 'cfg': cfg,
            'created': time.time(), 'finished': None,
        }
        with LOCK:
            TASKS[tid] = task

        def run():
            try:
                def prog(n):
                    with LOCK:
                        task['n_rows'] = n
                mapp = UPLOADS.get(sid, {}).get('mapping') if sid else None
                mp = load_mapping(mapp) if mapp else None
                h, n, prev = process_file(in_path, w3, w7, w14, out_path, cols=cols,
                                          on_progress=prog, cfg=cfg, mapping=mp)
                with LOCK:
                    task['status'] = 'done'
                    task['n_rows'] = n
                    task['header'] = h
                    task['preview_rows'] = prev
                    task['finished'] = time.time()
                try:
                    os.unlink(in_path)
                except OSError:
                    pass
                if mapp:
                    try:
                        os.unlink(mapp)
                    except OSError:
                        pass
            except Exception as e:
                import traceback
                with LOCK:
                    task['status'] = 'error'
                    task['error_msg'] = str(e)
                    task['traceback'] = traceback.format_exc()
                    task['finished'] = time.time()

        threading.Thread(target=run, daemon=True).start()
        self._json(202, {'id': tid, 'status': 'running', 'n_rows': 0})


def qs_get(p, key):
    v = urllib.parse.parse_qs(p.query).get(key)
    return v[0] if v else ''


def qs_get_list(p, key):
    return urllib.parse.parse_qs(p.query).get(key, [])


def _parse_cfg_param(p):
    """从 query 取 cfg(JSON, URL 编码)。无 cfg→None；非法→False；正常→dict。"""
    raw = qs_get(p, 'cfg')
    if not raw:
        return None
    try:
        return json.loads(urllib.parse.unquote(raw))
    except Exception:
        return False


def run_fba(sid, task):
    try:
        up = UPLOADS.get(sid, {})
        sales_p = up.get('sales'); ops_p = up.get('ops'); fba_p = up.get('fba')
        if not (sales_p and ops_p and fba_p):
            raise ValueError("缺少上传文件（销量 / 运营级别 / FBA需求）")
        tid = task['id']
        out_sales = os.path.join(HERE, '_fba_%s_sales.xlsx' % tid)
        out_ops = os.path.join(HERE, '_fba_%s_ops.xlsx' % tid)
        out_diff = os.path.join(HERE, '_fba_%s_diff.xlsx' % tid)

        def prog(n):
            with LOCK:
                task['n_rows'] = n

        # 运营级别映射表（slot=mapping）→ 备货规则按运营级别算加权日均
        mapp = up.get('mapping')
        mp = load_mapping(mapp) if mapp else None
        cfg = task.get('cfg')

        hdr_s, n_s, prev = process_file(sales_p, task['w'][0], task['w'][1], task['w'][2],
                                         out_sales, cols=task.get('cols'), on_progress=prog,
                                         cfg=cfg, mapping=mp)
        hdr_o, n_o, meet, level_info = process_ops(ops_p, out_sales, out_ops, cols=task.get('cols'),
                                                    on_progress=prog, levels=task.get('levels'), cfg=cfg, mapping=mp)
        hdr_d, n_d, summary, diff_preview = process_diff(out_ops, fba_p, out_diff, cols=task.get('cols'),
                                                         cfg=cfg, mapping=mp)

        with LOCK:
            task['status'] = 'done'
            task['outputs'] = {'sales': out_sales, 'ops': out_ops, 'diff': out_diff}
            task['summary'] = summary
            task['n_rows'] = n_s
            task['header'] = hdr_s
            task['preview_rows'] = prev
            task['diff_header'] = hdr_d
            task['diff_preview'] = diff_preview
            task['level_info'] = level_info
            task['finished'] = time.time()
        # 清理暂存文件
        for k in ('sales', 'ops', 'fba', 'mapping'):
            pp = up.get(k)
            if pp:
                try:
                    os.unlink(pp)
                except OSError:
                    pass
        UPLOADS.pop(sid, None)
    except Exception as e:
        import traceback
        with LOCK:
            task['status'] = 'error'
            task['error_msg'] = str(e)
            task['traceback'] = traceback.format_exc()
            task['finished'] = time.time()


def run_eu_check(sid, task):
    try:
        up = UPLOADS.get(sid, {})
        diff_p = up.get('diff'); sales_p = up.get('sales')
        if not (diff_p and sales_p):
            raise ValueError("缺少上传文件（差异表 / 销量原始）")
        tid = task['id']
        out_path = os.path.join(HERE, '_eu_%s.xlsx' % tid)
        w3, w7, w14 = task['w']

        def prog(n):
            with LOCK:
                task['n_rows'] = n

        hdr, n, summary, preview = process_eu_check(diff_p, sales_p, out_path, w3, w7, w14, on_progress=prog)
        with LOCK:
            task['status'] = 'done'
            task['output_path'] = out_path
            task['header'] = hdr
            task['preview_rows'] = preview
            task['summary'] = summary
            task['n_rows'] = n
            task['finished'] = time.time()
        for k in ('diff', 'sales'):
            pp = up.get(k)
            if pp:
                try:
                    os.unlink(pp)
                except OSError:
                    pass
        UPLOADS.pop(sid, None)
    except Exception as e:
        import traceback
        with LOCK:
            task['status'] = 'error'
            task['error_msg'] = str(e)
            task['traceback'] = traceback.format_exc()
            task['finished'] = time.time()


def run_eu_detail(sid, task):
    try:
        up = UPLOADS.get(sid, {})
        diff_p = up.get('diff'); sales_p = up.get('sales')
        if not (diff_p and sales_p):
            raise ValueError("缺少上传文件（差异表 / 销量原始）")
        tid = task['id']
        out_path = os.path.join(HERE, '_eu_detail_%s.xlsx' % tid)
        w3, w7, w14 = task['w']

        def prog(n):
            with LOCK:
                task['n_rows'] = n

        hdr, n_mismatch, summary = process_eu_mismatch_detail(diff_p, sales_p, out_path, w3, w7, w14, on_progress=prog)
        with LOCK:
            task['status'] = 'done'
            task['output_path'] = out_path
            task['header'] = hdr
            task['summary'] = summary
            task['n_rows'] = n_mismatch
            task['finished'] = time.time()
        for k in ('diff', 'sales'):
            pp = up.get(k)
            if pp:
                try:
                    os.unlink(pp)
                except OSError:
                    pass
        UPLOADS.pop(sid, None)
    except Exception as e:
        import traceback
        with LOCK:
            task['status'] = 'error'
            task['error_msg'] = str(e)
            task['traceback'] = traceback.format_exc()
            task['finished'] = time.time()


class ThreadingServer(http.server.ThreadingHTTPServer):
    allow_reuse_address = True


if __name__ == '__main__':
    print("服务已启动： http://localhost:%d  （Ctrl+C 停止）" % PORT)
    # 自检：打印当前代码实际包含的关键功能。重启后看这一行，
    # 如果缺少 +FBA 上传 或 +FBA 任务，说明跑的还是旧版 server.py。
    extras = []
    if hasattr(H, 'do_OPTIONS'):
        extras.append('OPTIONS预检')
    try:
        src = inspect.getsource(H.do_POST)
        if '/api/upload' in src:
            extras.append('FBA上传')
        if 'run_fba' in src:
            extras.append('FBA任务')
        if 'process_eu_check' in src:
            extras.append('EU核对')
        if 'process_eu_mismatch_detail' in src:
            extras.append('EU明细')
        if 'process_alloc_check' in src:
            extras.append('分配核对')
        if 'append_types' in src and 'is_append_row' in src:
            extras.append('追加识别')
    except Exception:
        pass
    print("当前版本： %s" % ('基础处理' if not extras else '基础处理 + ' + ' + '.join(extras)))
    ThreadingServer(('0.0.0.0', PORT), H).serve_forever()
