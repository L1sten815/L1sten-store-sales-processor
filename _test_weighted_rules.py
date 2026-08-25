"""加权日均销量（按运营级别备货规则）回归测试。
覆盖：legacy 模式、rules 模式（组合/固定/未绑定 fallback/前14天=0但30d有销量）、页签② 联动、缺列容错。
运行：venv python _test_weighted_rules.py
"""
import os, sys, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
import server

HERE = tempfile.mkdtemp()
HDR = ['店铺','ASIN','状态','三天日均销量','七天日均销量','十四天日均销量',
       '三十天日均销量','60天日均销量','九十天日均销量']

def mk_sales(rows, path=None):
    path = path or os.path.join(HERE, 'sales_in.xlsx')
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(HDR)
    for r in rows: ws.append(r)
    wb.save(path)
    return path

def M(prev, ri, h):
    return prev[ri][h.index('日均销量加权')]

ROWS = [
    ['S1','A1','正常销售', 4, 3, 2, 1, 0.5, 0.2],   # 新品 → 30天测试
    ['S2','A2','正常销售', 0, 0, 0, 0, 5.0, 0],       # 美国-30天日均-[5~10] → 固定60
    ['S3','A3','正常销售', 1, 1, 1, 1, 1, 1],         # 未绑定 → default 30天测试
    ['S4','A4','正常销售', 0, 0, 0, 2.0, 2.5, 3.0],   # 新品 → 前14天=0 但30d有销量
    ['S5','A5','正常销售', 0, 0, 0, 0, 0, 0],         # 新品 → 全0
]
sales_in = mk_sales(ROWS)

CFG = {
    "mode": "rules",
    "rules": [
        {"name":"30天测试","type":"组合","weights":{"3":0.5,"7":0.2,"14":0.2,"30":0.1,"60":0,"90":0}},
        {"name":"60天测试日均","type":"固定","fixed_window":60},
    ],
    "bindings": {"新品":"30天测试","美国-30天日均-[5~10]":"60天测试日均"},
    "default_rule": "30天测试",
    "default_op_level": "",
}
MAPPING = {"S1A1":("新品",""), "S2A2":("美国-30天日均-[5~10]",""), "S4A4":("新品",""), "S5A5":("新品","")}

fails = []
def check(name, got, exp):
    try:
        got = float(got)
    except Exception:
        pass
    ok = isinstance(got, (int, float)) and abs(got - exp) < 1e-6
    print(f"  [{'✅' if ok else '❌'}] {name}: 得到 {got}  期望 {exp}")
    if not ok: fails.append(name)

print("=== 1) legacy 模式（0.6/0.3/0.1）===")
legacy_out = os.path.join(HERE, 'legacy_out.xlsx')
h, n, prev = server.process_file(sales_in, 0.6, 0.3, 0.1, legacy_out)
check("legacy S1", M(prev, 0, h), 3.5)   # 0.6*4+0.3*3+0.1*2

print("\n=== 2) rules 模式 + 映射 ===")
rules_out = os.path.join(HERE, 'rules_out.xlsx')
h, n, prev = server.process_file(sales_in, 0.6, 0.3, 0.1, rules_out, cfg=CFG, mapping=MAPPING)
check("组合 新品 S1",     M(prev, 0, h), 0.5*4+0.2*3+0.2*2+0.1*1)   # 3.1
check("固定 60窗 S2",     M(prev, 1, h), 5.0)
check("未绑定→default S3", M(prev, 2, h), 0.5*1+0.2*1+0.2*1+0.1*1) # 1.0
check("前14天=0 但30d有销量 S4", M(prev, 3, h), 0.1*2.0)           # 0.2 >0
check("全0 S5",          M(prev, 4, h), 0.0)

print("\n=== 3) 页签② 联动：前14天=0 但30d有销量 → 满足条件=是 ===")
ops_in = os.path.join(HERE, 'ops_in.xlsx')
wb = openpyxl.Workbook(); ws = wb.active
ws.append(['店铺ASIN','销售状态','配送类型','运营级别'])
ws.append(['S4A4','正常销售','FBA','新品'])   # 销量>0 → 应生成
ws.append(['S5A5','正常销售','FBA','新品'])   # 销量=0 → 不生成
wb.save(ops_in)
ops_out = os.path.join(HERE, 'ops_out.xlsx')
# 页签②已改为直接读销量原始文件 lookup，不再传 process_file 处理后的结果
oh, on, meet, level_info = server.process_ops(ops_in, sales_in, ops_out, 0.6, 0.3, 0.1, cfg=CFG, mapping=MAPPING)
wb2 = openpyxl.load_workbook(ops_out, data_only=True); ws2 = wb2.active
rows = list(ws2.iter_rows(values_only=True))
hdr = list(rows[0]); ci = hdr.index('满足条件')
vals = [r[ci] for r in rows[1:]]
print(f"  meet={meet}  满足条件列={vals}")
check("页签② 联动 meet", meet, 1)
if vals != ['是','否']:
    fails.append("页签② 满足条件值应为 [是,否]")

print("\n=== 4) 缺列容错：文件无 30/60/90 列时按 0（仍正常）===")
sales_short = mk_sales([['X1','X2','正常销售',2,2,2]], os.path.join(HERE,'short.xlsx'))
so = os.path.join(HERE,'short_out.xlsx')
h, n, prev = server.process_file(sales_short, 0.6,0.3,0.1, so, cfg=CFG, mapping={"X1X2":("新品","")})
# 新品→30天测试，无30/60/90列→0：0.5*2+0.2*2+0.2*2 = 1.8
check("缺窗列容错 X1", M(prev, 0, h), 1.8)

print("\n=== 结果 ===")
if fails:
    print("❌ 失败用例:", fails); sys.exit(1)
print("✅ 全部通过")
